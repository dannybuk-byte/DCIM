"""
NYISO interconnection queue adapter (XLSX ingest).
URL discovery parses https://www.nyiso.com/interconnections HTML; tests use fixtures only.
"""

from __future__ import annotations

import re
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

from ..models import QueueRecord
from ..normalize import geography_from_location_raw, normalize_developer_name
from ..queue_normalize import infer_queue_type, normalize_status
from .base import QueueAdapter

NYISO_INTERCONNECTIONS_PAGE = "https://www.nyiso.com/interconnections"
NYISO_DEFAULT_XLSX = (
    "https://www.nyiso.com/documents/20142/1407078/NYISO-Interconnection-Queue.xlsx"
)

_XLSX_HREF = re.compile(
    r'href="([^"]*NYISO-Interconnection-Queue\.xlsx[^"]*)"',
    re.IGNORECASE,
)


def discover_nyiso_queue_xlsx_url(html: str, base_url: str = NYISO_INTERCONNECTIONS_PAGE) -> Optional[str]:
    m = _XLSX_HREF.search(html)
    if not m:
        return None
    href = m.group(1)
    if href.startswith("http"):
        return href
    return urljoin(base_url, href)


class NyisoQueueAdapter(QueueAdapter):
    iso = "NYISO"

    def __init__(
        self,
        *,
        xlsx_path: Optional[Path] = None,
        xlsx_url: Optional[str] = None,
        discover_url: bool = False,
    ) -> None:
        self.xlsx_path = xlsx_path
        self.xlsx_url = xlsx_url
        self.discover_url = discover_url

    def fetch_queue_records(self) -> list[QueueRecord]:
        if self.xlsx_path and self.xlsx_path.is_file():
            data = self.xlsx_path.read_bytes()
            return self._parse_xlsx(data, source_url=str(self.xlsx_path))
        url = self.xlsx_url
        if self.discover_url and not url:
            url = self._discover_url() or NYISO_DEFAULT_XLSX
        if not url:
            return []
        resp = requests.get(url, timeout=120)
        resp.raise_for_status()
        return self._parse_xlsx(resp.content, source_url=url)

    def _discover_url(self) -> Optional[str]:
        resp = requests.get(NYISO_INTERCONNECTIONS_PAGE, timeout=60)
        resp.raise_for_status()
        return discover_nyiso_queue_xlsx_url(resp.text)

    def _parse_xlsx(self, data: bytes, source_url: str) -> list[QueueRecord]:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(c or "").strip().lower() for c in next(rows_iter, [])]
        col = {name: i for i, name in enumerate(header)}

        def cell(row: tuple, *names: str) -> str:
            for n in names:
                idx = col.get(n)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
            return ""

        fetched = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        out: list[QueueRecord] = []
        for row in rows_iter:
            if not row or not any(row):
                continue
            project = cell(row, "project name", "queue position", "name")
            developer = cell(row, "interconnection customer", "owner", "developer")
            if not developer:
                developer = project
            queue_id = cell(row, "queue position", "queue pos.", "qpt") or project
            status_raw = cell(row, "status", "queue status") or "unknown"
            status_norm = normalize_status(status_raw)
            location_raw = cell(row, "county", "state", "location") or ""
            cap_raw = cell(row, "mw", "capacity", "mw capacity")
            capacity = None
            if cap_raw:
                try:
                    capacity = float(re.sub(r"[^\d.]", "", cap_raw))
                except ValueError:
                    capacity = None
            raw_dict = {
                "project_type": cell(row, "type", "fuel", "technology"),
                "technology": cell(row, "technology", "fuel"),
            }
            qtype = infer_queue_type(raw_dict, self.iso)
            signal_class = "planning"
            if status_norm == "construction":
                signal_class = "construction"
            elif status_norm == "operational":
                signal_class = "activation"
            loc_norm = geography_from_location_raw(location_raw)
            out.append(
                QueueRecord(
                    queue_id=queue_id or f"NYISO-{len(out)}",
                    iso=self.iso,
                    project_name=project or developer,
                    developer_name=developer,
                    developer_name_normalized=normalize_developer_name(developer),
                    queue_type=qtype,  # type: ignore[arg-type]
                    capacity_mw=capacity,
                    proposed_in_service_date=cell(row, "proposed in-service", "in service") or None,
                    status=status_raw,
                    status_normalized=status_norm,  # type: ignore[arg-type]
                    location_raw=location_raw,
                    location_normalized=loc_norm,
                    fetched_at=fetched,
                    source_url=source_url,
                    signal_class=signal_class,  # type: ignore[arg-type]
                )
            )
        wb.close()
        return out
